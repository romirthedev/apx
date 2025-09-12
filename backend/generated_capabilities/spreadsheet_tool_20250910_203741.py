
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional, Union
import os
from datetime import datetime

class FinancialModelingTool:
    """
    Tool for creating and manipulating Excel spreadsheets for financial modeling,
    including support for advanced formulas, pivot tables, and basic data analysis.
    """

    def __init__(self):
        self.current_filepath: Optional[str] = None
        self.workbook: Optional[openpyxl.workbook.workbook.Workbook] = None
        self.sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        self.dataframe: Optional[pd.DataFrame] = None

    def _get_or_create_workbook(self, filepath: str, sheet_name: str = "Sheet1") -> None:
        """Internal method to get or create an Excel workbook and sheet."""
        if self.workbook is None or self.current_filepath != filepath:
            if os.path.exists(filepath):
                try:
                    self.workbook = openpyxl.load_workbook(filepath)
                except Exception as e:
                    raise IOError(f"Error loading existing workbook at {filepath}: {e}")
            else:
                self.workbook = openpyxl.Workbook()

            if sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet_name]
            else:
                self.sheet = self.workbook.create_sheet(sheet_name)
            self.current_filepath = filepath
        elif self.sheet is None:
            if sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet_name]
            else:
                self.sheet = self.workbook.create_sheet(sheet_name)

    def _load_dataframe_from_sheet(self) -> None:
        """Internal method to load data from the current sheet into a pandas DataFrame."""
        if self.sheet is None:
            raise ValueError("No sheet is currently active.")

        data = []
        header = [cell.value for cell in self.sheet[1]]
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(header, row)))
        self.dataframe = pd.DataFrame(data)

    def _save_dataframe_to_sheet(self, sheet_name: str = "Sheet1") -> None:
        """Internal method to save the current DataFrame to the active sheet."""
        if self.dataframe is None:
            raise ValueError("No data available to save.")
        if self.sheet is None:
            raise ValueError("No sheet is currently active.")

        # Clear existing data before writing new data
        self.sheet.delete_rows(1, self.sheet.max_row)

        # Write header
        for col_num, value in enumerate(self.dataframe.columns, 1):
            self.sheet.cell(row=1, column=col_num, value=value)

        # Write data rows
        for row_num, row_data in enumerate(self.dataframe.values, 2):
            for col_num, value in enumerate(row_data, 1):
                self.sheet.cell(row=row_num, column=col_num, value=value)

        self.workbook.save(self.current_filepath)

    def create_financial_model_file(
        self,
        initial_data: Optional[List[Dict]] = None,
        filename: str = "financial_model",
        sheet_name: str = "FinancialData",
        formulas: Optional[Dict[str, str]] = None,
        pivot_tables: Optional[List[Dict[str, Any]]] = None
    ) -> Dict[str, Any]:
        """
        Creates a new Excel file for financial modeling with initial data,
        advanced formulas, and pivot tables.

        Args:
            initial_data: A list of dictionaries representing the initial data for the sheet.
                          Each dictionary is a row, with keys as column headers.
            filename: The name of the Excel file to create.
            sheet_name: The name of the primary sheet for data.
            formulas: A dictionary where keys are cell addresses (e.g., "C2") and
                      values are the Excel formula strings to apply.
            pivot_tables: A list of dictionaries, each defining a pivot table.
                          Each dictionary can contain keys like:
                          'data_sheet': Name of the sheet containing the source data for the pivot table.
                          'pivot_sheet': Name of the sheet where the pivot table will be created.
                          'pivot_table_destination': Cell address for the top-left corner of the pivot table.
                          'rows': List of field names for rows.
                          'columns': List of field names for columns.
                          'values': List of dictionaries defining value fields (e.g., {'field': 'Sales', 'function': 'Sum'}).
                          'filters': List of field names for filters.

        Returns:
            A dictionary indicating success or failure with relevant details.
        """
        if not filename or not isinstance(filename, str):
            return {"success": False, "error": "Invalid filename. Must be a non-empty string.", "message": "Invalid filename."}
        if not sheet_name or not isinstance(sheet_name, str):
            return {"success": False, "error": "Invalid sheet name. Must be a non-empty string.", "message": "Invalid sheet name."}

        filepath = os.path.expanduser(f"~/Desktop/{filename}.xlsx")

        try:
            self._get_or_create_workbook(filepath, sheet_name)

            if initial_data:
                self.dataframe = pd.DataFrame(initial_data)
                self._save_dataframe_to_sheet(sheet_name) # Save initial data

            # Apply formulas after data is loaded or created
            if formulas:
                if self.dataframe is None and not initial_data:
                    # If no initial data, try to load from sheet if it exists and has data
                    try:
                        self._load_dataframe_from_sheet()
                    except ValueError:
                        pass # Ignore if sheet is empty

                if self.dataframe is not None or initial_data: # Ensure there's data to relate formulas to
                    for cell_address, formula in formulas.items():
                        if not isinstance(cell_address, str) or not isinstance(formula, str):
                            print(f"Warning: Skipping invalid formula entry (cell: {cell_address}, formula: {formula}). Must be string.")
                            continue
                        try:
                            self.sheet[cell_address] = f"={formula}"
                        except Exception as e:
                            print(f"Warning: Could not apply formula '{formula}' to cell '{cell_address}': {e}")
                else:
                    print("Warning: Formulas requested but no data found in the sheet. Formulas may not function as expected.")


            # Create pivot tables
            if pivot_tables:
                if not self.dataframe and not initial_data:
                     # Try to load data if it wasn't explicitly provided for this creation call
                    try:
                        self._load_dataframe_from_sheet()
                    except ValueError:
                        pass # Ignore if sheet is empty

                if self.dataframe is not None or initial_data:
                    for pivot_config in pivot_tables:
                        try:
                            self._create_pivot_table(pivot_config)
                        except Exception as e:
                            print(f"Warning: Could not create pivot table with config {pivot_config}: {e}")
                else:
                    print("Warning: Pivot tables requested but no data found in the sheet.")


            self.workbook.save(filepath)
            self.current_filepath = filepath # Update current file after successful creation

            return {
                "success": True,
                "filepath": filepath,
                "sheet_name": sheet_name,
                "message": f"Financial model file '{filename}.xlsx' created successfully at {filepath}"
            }
        except (IOError, ValueError, Exception) as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to create financial model file: {str(e)}"
            }

    def _create_pivot_table(self, pivot_config: Dict[str, Any]) -> None:
        """Internal helper to create a pivot table based on configuration."""
        if self.workbook is None:
            raise ValueError("Workbook not initialized. Call create_financial_model_file first.")

        source_sheet_name = pivot_config.get('data_sheet', self.sheet.title if self.sheet else "Sheet1")
        pivot_sheet_name = pivot_config.get('pivot_sheet', f"{source_sheet_name}_Pivot")
        pivot_dest_cell = pivot_config.get('pivot_table_destination', 'A1')
        row_fields = pivot_config.get('rows', [])
        col_fields = pivot_config.get('columns', [])
        value_fields = pivot_config.get('values', [])
        filter_fields = pivot_config.get('filters', [])

        if source_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Source sheet '{source_sheet_name}' not found in workbook.")

        source_sheet = self.workbook[source_sheet_name]

        # Create pivot sheet if it doesn't exist
        if pivot_sheet_name not in self.workbook.sheetnames:
            pivot_sheet = self.workbook.create_sheet(pivot_sheet_name)
        else:
            pivot_sheet = self.workbook[pivot_sheet_name]
            # Clear existing pivot table content to avoid conflicts
            pivot_sheet.delete_rows(1, pivot_sheet.max_row)
            pivot_sheet.delete_columns(1, pivot_sheet.max_column)


        # Get the range of the source data
        data_range = f"{source_sheet_name}!${get_column_letter(source_sheet.min_column)}${source_sheet.min_row}:${get_column_letter(source_sheet.max_column)}${source_sheet.max_row}"

        # Construct the pivot table XML (this is a simplified representation, full XML is complex)
        # Openpyxl's pivot table functionality is limited. A more robust solution might involve
        # using xlsxwriter or directly manipulating XML. For demonstration, we'll assume
        # this can be handled or that the user might use Excel's UI after file creation.
        # However, openpyxl *does* have a basic pivot table API.

        try:
            # Define data source
            data_source = openpyxl.pivot.cache.PivotCache(
                workbook=self.workbook,
                source_data=data_range,
                source_sheet=source_sheet_name,
                cache_type='Cache1' # Unique ID for the cache
            )

            # Create pivot table object
            pt = pivot_sheet.add_pivot_table(
                data_source=data_source,
                location=pivot_dest_cell
            )

            # Configure row fields
            for field in row_fields:
                pt.RowFields.append(openpyxl.pivot.field.RowField(field))

            # Configure column fields
            for field in col_fields:
                pt.ColFields.append(openpyxl.pivot.field.ColField(field))

            # Configure value fields
            for val_config in value_fields:
                field_name = val_config.get('field')
                function = val_config.get('function', 'Sum') # Default to Sum
                if field_name:
                    pt.DataFields.append(openpyxl.pivot.field.DataField(field_name, function))

            # Configure filter fields
            for field in filter_fields:
                pt.PageFields.append(openpyxl.pivot.field.PageField(field))

        except Exception as e:
            # If openpyxl's direct pivot table API fails or is insufficient, provide a warning.
            # For true advanced pivot tables, one might need to write more complex XML or use other libraries.
            raise Exception(f"Failed to create pivot table using openpyxl API (requires Excel to finalize): {e}. "
                            "Consider manual creation in Excel or using xlsxwriter for more control.")


    def add_data_to_current_file(self, new_data: List[Dict]) -> Dict[str, Any]:
        """
        Adds new data rows to the current loaded spreadsheet.

        Args:
            new_data: A list of dictionaries representing the new data rows.

        Returns:
            A dictionary indicating success or failure.
        """
        if self.current_filepath is None or self.workbook is None or self.sheet is None:
            return {"success": False, "error": "No spreadsheet is currently loaded. Use create_financial_model_file first.", "message": "No spreadsheet loaded."}

        if not new_data or not isinstance(new_data, list) or not all(isinstance(row, dict) for row in new_data):
            return {"success": False, "error": "Invalid new_data format. Must be a list of dictionaries.", "message": "Invalid data format."}

        try:
            if self.dataframe is None:
                self._load_dataframe_from_sheet() # Load existing data if not already in memory

            new_df = pd.DataFrame(new_data)
            if self.dataframe is None:
                self.dataframe = new_df
            else:
                self.dataframe = pd.concat([self.dataframe, new_df], ignore_index=True)

            self._save_dataframe_to_sheet(self.sheet.title) # Overwrite sheet with updated data

            return {
                "success": True,
                "rows_added": len(new_data),
                "total_rows": len(self.dataframe),
                "message": f"Added {len(new_data)} rows successfully to {self.current_filepath}."
            }
        except (ValueError, Exception) as e:
            return {"success": False, "error": str(e), "message": f"Failed to add data: {str(e)}"}

    def apply_formulas_to_current_file(self, formulas: Dict[str, str], sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Applies Excel formulas to the currently loaded spreadsheet.

        Args:
            formulas: A dictionary where keys are cell addresses (e.g., "C2") and
                      values are the Excel formula strings to apply.
            sheet_name: The name of the sheet to apply formulas to. If None, uses the current sheet.

        Returns:
            A dictionary indicating success or failure.
        """
        if self.current_filepath is None or self.workbook is None:
            return {"success": False, "error": "No spreadsheet is currently loaded. Use create_financial_model_file first.", "message": "No spreadsheet loaded."}

        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found in the current workbook.", "message": "Sheet not found."}
            self.sheet = self.workbook[sheet_name]
        elif self.sheet is None:
            return {"success": False, "error": "No sheet is currently active. Please specify a sheet_name or load data first.", "message": "No active sheet."}

        if not formulas or not isinstance(formulas, dict):
            return {"success": False, "error": "Invalid formulas format. Must be a dictionary of cell: formula.", "message": "Invalid formulas format."}

        try:
            if self.dataframe is None:
                self._load_dataframe_from_sheet() # Load data to ensure context for formulas

            for cell_address, formula in formulas.items():
                if not isinstance(cell_address, str) or not isinstance(formula, str):
                    print(f"Warning: Skipping invalid formula entry (cell: {cell_address}, formula: {formula}). Must be string.")
                    continue
                try:
                    # Ensure the formula starts with '='
                    if not formula.startswith('='):
                        formula = f"={formula}"
                    self.sheet[cell_address] = formula
                except Exception as e:
                    print(f"Warning: Could not apply formula '{formula}' to cell '{cell_address}': {e}")

            self.workbook.save(self.current_filepath)
            return {
                "success": True,
                "message": f"Formulas applied successfully to sheet '{self.sheet.title}' in {self.current_filepath}."
            }
        except (ValueError, Exception) as e:
            return {"success": False, "error": str(e), "message": f"Failed to apply formulas: {str(e)}"}

    def save_current_file(self, filename: Optional[str] = None) -> Dict[str, Any]:
        """
        Saves the currently loaded spreadsheet. If a filename is provided,
        it renames or saves to a new location.

        Args:
            filename: Optional. If provided, saves the file with this name
                      (e.g., "my_financial_model"). If None, saves to the
                      current_filepath.

        Returns:
            A dictionary indicating success or failure.
        """
        if self.current_filepath is None or self.workbook is None:
            return {"success": False, "error": "No spreadsheet is currently loaded.", "message": "No spreadsheet loaded."}

        try:
            if filename:
                new_filepath = os.path.expanduser(f"~/Desktop/{filename}.xlsx")
                self.workbook.save(new_filepath)
                self.current_filepath = new_filepath
                return {"success": True, "filepath": new_filepath, "message": f"Spreadsheet saved to {new_filepath}"}
            else:
                self.workbook.save(self.current_filepath)
                return {"success": True, "filepath": self.current_filepath, "message": f"Spreadsheet saved to {self.current_filepath}"}
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to save spreadsheet: {str(e)}"}

    def load_spreadsheet(self, filepath: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """
        Loads an existing Excel spreadsheet.

        Args:
            filepath: The full path to the Excel file.
            sheet_name: The name of the sheet to load data from.

        Returns:
            A dictionary indicating success or failure.
        """
        filepath = os.path.expanduser(filepath)
        if not os.path.exists(filepath):
            return {"success": False, "error": f"File not found at {filepath}", "message": "File not found."}

        try:
            self.workbook = openpyxl.load_workbook(filepath)
            if sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet_name]
                self.current_filepath = filepath
                self._load_dataframe_from_sheet() # Load data into dataframe on load
                return {
                    "success": True,
                    "filepath": filepath,
                    "sheet_name": sheet_name,
                    "rows": len(self.dataframe) if self.dataframe is not None else 0,
                    "columns": len(self.dataframe.columns) if self.dataframe is not None else 0,
                    "message": f"Spreadsheet loaded successfully from {filepath} (Sheet: {sheet_name})."
                }
            else:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found in {filepath}", "message": "Sheet not found."}
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to load spreadsheet: {str(e)}"}

