
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from typing import Dict, Any, List, Optional
import os
from datetime import datetime

class FinancialDashboardTool:
    """
    Tool for creating and manipulating spreadsheets, specifically designed
    to assist in building a comprehensive financial dashboard in Excel.
    This tool leverages openpyxl to add capabilities beyond basic data
    writing, enabling the creation of automated pivot tables, ROI formulas,
    and dynamic charts, although direct programmatic configuration of these
    advanced Excel features (like VBA/Macros) is outside its scope.
    The output Excel file will be structured to facilitate manual setup
    of these advanced features by the user or by subsequent scripting.
    """

    def __init__(self):
        self.current_filepath = None
        self.workbook = None
        self.sheet = None
        self.data_sheet_name = "FinancialData"
        self.dashboard_sheet_name = "Dashboard"
        self.pivot_sheet_name = "PivotTables"
        self.charts_sheet_name = "Charts"

    def _get_or_create_workbook(self, filename: str, sheet_name: str = None) -> openpyxl.workbook.workbook.Workbook:
        """Helper to get or create a workbook and worksheet."""
        if self.workbook is None:
            try:
                self.current_filepath = os.path.expanduser(f"~/Desktop/{filename}")
                if os.path.exists(self.current_filepath):
                    self.workbook = openpyxl.load_workbook(self.current_filepath)
                else:
                    self.workbook = openpyxl.Workbook()
                
                # Ensure all required sheets exist
                sheets_to_ensure = [self.data_sheet_name, self.dashboard_sheet_name, self.pivot_sheet_name, self.charts_sheet_name]
                for s_name in sheets_to_ensure:
                    if s_name not in self.workbook.sheetnames:
                        self.workbook.create_sheet(s_name)
                
                # Set the active sheet for data writing
                if sheet_name:
                    if sheet_name in self.workbook.sheetnames:
                        self.sheet = self.workbook[sheet_name]
                    else:
                        self.sheet = self.workbook.create_sheet(sheet_name)
                else:
                    self.sheet = self.workbook[self.data_sheet_name]

            except Exception as e:
                raise RuntimeError(f"Failed to initialize workbook: {e}")
        else:
            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    self.sheet = self.workbook[sheet_name]
                else:
                    self.sheet = self.workbook.create_sheet(sheet_name)
            else:
                self.sheet = self.workbook[self.data_sheet_name]
        return self.workbook

    def _validate_data(self, data: List[Dict]):
        """Validates the input data format."""
        if not isinstance(data, list):
            raise TypeError("Data must be a list of dictionaries.")
        if not data:
            return # Empty data is acceptable, but not ideal for a dashboard
        if not all(isinstance(item, dict) for item in data):
            raise TypeError("Each item in the data list must be a dictionary.")
        
        # Check for consistent keys if there's more than one item
        if len(data) > 1:
            keys = data[0].keys()
            if not all(item.keys() == keys for item in data):
                # Warn or raise, depending on strictness. For now, warn.
                print("Warning: Dictionaries in data have inconsistent keys.")

    def create_or_update_spreadsheet(self, data: List[Dict], filename: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        Creates a new Excel file or updates an existing one with provided data.
        It also sets up basic structure for dashboard elements.
        """
        try:
            self._validate_data(data)
            
            # Determine filename and initialize workbook
            current_filename_for_save = filename # Use provided filename for saving
            if not current_filename_for_save.lower().endswith('.xlsx'):
                current_filename_for_save += '.xlsx'
            
            self.current_filepath = os.path.expanduser(f"~/Desktop/{current_filename_for_save}")
            self.workbook = None # Reset workbook to ensure fresh load or creation
            
            # Get or create workbook, ensuring necessary sheets exist
            self.workbook = self._get_or_create_workbook(current_filename_for_save, sheet_name=sheet_name)
            
            # If a specific sheet_name was provided, use it. Otherwise default to data sheet.
            target_sheet = self.sheet if sheet_name else self.workbook[self.data_sheet_name]
            target_sheet.title = sheet_name if sheet_name else self.data_sheet_name # Ensure title is correct

            # Clear previous data in the target sheet if it's not a new file or a new sheet
            if target_sheet.max_row > 0:
                for row in list(target_sheet.rows):
                    for cell in row:
                        cell.value = None

            # Write data
            if data:
                df = pd.DataFrame(data)
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        target_sheet.cell(row=r_idx, column=c_idx, value=value)
                
                rows = len(df)
                columns = len(df.columns)
            else:
                rows = 0
                columns = 0
                target_sheet.cell(row=1, column=1, value="No data provided.")

            # Save the workbook
            self.workbook.save(self.current_filepath)

            return {
                "success": True,
                "filepath": self.current_filepath,
                "sheet_name": target_sheet.title,
                "rows": rows,
                "columns": columns,
                "message": f"Spreadsheet '{current_filename_for_save}' created/updated successfully at {self.current_filepath}."
            }
        except (TypeError, ValueError) as ve:
            return {
                "success": False,
                "error": str(ve),
                "message": f"Data validation failed: {str(ve)}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to create/update spreadsheet: {str(e)}"
            }

    def add_data_to_sheet(self, new_data: List[Dict], sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Adds new data to an existing sheet in the current workbook.
        If no workbook is loaded, it will attempt to load the last known file.
        """
        try:
            if self.workbook is None and self.current_filepath is None:
                return {"success": False, "error": "No spreadsheet loaded or created yet. Use create_or_update_spreadsheet first."}
            
            if self.workbook is None: # Attempt to load last known file
                if not os.path.exists(self.current_filepath):
                    return {"success": False, "error": f"Last known file not found: {self.current_filepath}"}
                self.workbook = openpyxl.load_workbook(self.current_filepath)
            
            self._validate_data(new_data)
            
            target_sheet_name = sheet_name if sheet_name else self.data_sheet_name
            if target_sheet_name not in self.workbook.sheetnames:
                self.sheet = self.workbook.create_sheet(target_sheet_name)
            else:
                self.sheet = self.workbook[target_sheet_name]

            # Find the last row with data to append
            last_row = self.sheet.max_row
            start_row = last_row + 1 if last_row > 0 else 1

            if not new_data:
                return {"success": True, "rows_added": 0, "total_rows": last_row, "message": "No new data to add."}

            # Assuming new_data has same columns as existing data or we're adding to an empty sheet
            # For robustness, one might check column consistency or prompt user.
            df_new = pd.DataFrame(new_data)
            
            # Write new data, starting from the next available row
            for r_idx, row in enumerate(dataframe_to_rows(df_new, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    self.sheet.cell(row=r_idx, column=c_idx, value=value)

            current_max_row = self.sheet.max_row
            self.workbook.save(self.current_filepath)

            return {
                "success": True,
                "rows_added": len(new_data),
                "total_rows": current_max_row,
                "sheet_name": target_sheet_name,
                "message": f"Added {len(new_data)} rows to sheet '{target_sheet_name}'."
            }
        except (TypeError, ValueError) as ve:
            return {
                "success": False,
                "error": str(ve),
                "message": f"Data validation failed: {str(ve)}"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to add data: {str(e)}"
            }
    
    def save_workbook(self, filepath: str = None) -> Dict[str, Any]:
        """Saves the current workbook to a specified or default path."""
        try:
            if self.workbook is None:
                return {"success": False, "error": "No workbook is currently loaded or created."}
            
            save_path = filepath or self.current_filepath
            if not save_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                save_path = os.path.expanduser(f"~/Desktop/financial_dashboard_{timestamp}.xlsx")
            
            if not save_path.lower().endswith('.xlsx'):
                save_path += '.xlsx'

            self.workbook.save(save_path)
            self.current_filepath = save_path # Update current file path

            return {
                "success": True,
                "filepath": save_path,
                "message": f"Workbook saved successfully to {save_path}"
            }
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to save workbook: {str(e)}"}

    def setup_roi_calculations(self, 
                                sheet_name: str = "Dashboard", 
                                data_source_sheet: str = "FinancialData", 
                                investment_col: str = "Investment", 
                                revenue_col: str = "Revenue", 
                                date_col: str = "Date",
                                roi_col_name: str = "ROI (%)",
                                net_profit_col_name: str = "Net Profit") -> Dict[str, Any]:
        """
        Adds formulas for ROI and Net Profit calculations to a specified sheet.
        These formulas will reference data from another sheet.
        Note: This function adds formulas, not VBA macros. The user will need
        to enable macros or manually update if they want true automation.
        """
        try:
            if self.workbook is None:
                return {"success": False, "error": "No workbook loaded. Please create or load a spreadsheet first."}

            if sheet_name not in self.workbook.sheetnames:
                self.sheet = self.workbook.create_sheet(sheet_name)
            else:
                self.sheet = self.workbook[sheet_name]
            
            # Ensure data source sheet and columns exist
            if data_source_sheet not in self.workbook.sheetnames:
                return {"success": False, "error": f"Data source sheet '{data_source_sheet}' not found."}
            
            data_sheet = self.workbook[data_source_sheet]
            if investment_col not in data_sheet[1]: # Check header row
                 return {"success": False, "error": f"Investment column '{investment_col}' not found in '{data_source_sheet}'."}
            if revenue_col not in data_sheet[1]:
                 return {"success": False, "error": f"Revenue column '{revenue_col}' not found in '{data_source_sheet}'."}
            if date_col not in data_sheet[1]:
                 return {"success": False, "error": f"Date column '{date_col}' not found in '{data_source_sheet}'."}

            # Find column indices in the data source sheet
            header_row = [cell.value for cell in data_sheet[1]]
            try:
                inv_col_idx = header_row.index(investment_col) + 1
                rev_col_idx = header_row.index(revenue_col) + 1
                date_col_idx = header_row.index(date_col) + 1
            except ValueError as e:
                return {"success": False, "error": f"Column not found in data sheet header: {e}"}

            # Add headers to the dashboard sheet
            dashboard_header_row = [roi_col_name, net_profit_col_name]
            for c_idx, header_text in enumerate(dashboard_header_row, start=1):
                cell = self.sheet.cell(row=1, column=c_idx)
                cell.value = header_text
                cell.font = Font(bold=True)

            # Add formulas
            num_data_rows = data_sheet.max_row - 1 # Exclude header
            if num_data_rows > 0:
                # ROI Formula: ((Revenue - Investment) / Investment) * 100
                roi_formula = f'=IFERROR(((\'{data_source_sheet}\'!${chr(ord("A") + rev_col_idx - 1)}{row_num} - \'{data_source_sheet}\'!${chr(ord("A") + inv_col_idx - 1)}{row_num}) / \'{data_source_sheet}\'!${chr(ord("A") + inv_col_idx - 1)}{row_num}), 0)'
                # Net Profit Formula: Revenue - Investment
                net_profit_formula = f'=IFERROR(\'{data_source_sheet}\'!${chr(ord("A") + rev_col_idx - 1)}{row_num} - \'{data_source_sheet}\'!${chr(ord("A") + inv_col_idx - 1)}{row_num}, 0)'

                for i in range(2, num_data_rows + 2): # Start from row 2, including header
                    # ROI Calculation cell
                    roi_cell = self.sheet.cell(row=i, column=1) # Column 1 for ROI
                    roi_cell.value = roi_formula.replace("{row_num}", str(i-1)) # Adjust for 0-based index in formula string for data sheet
                    
                    # Net Profit Calculation cell
                    net_profit_cell = self.sheet.cell(row=i, column=2) # Column 2 for Net Profit
                    net_profit_cell.value = net_profit_formula.replace("{row_num}", str(i-1)) # Adjust

                # Apply percentage formatting for ROI
                for i in range(2, num_data_rows + 2):
                    self.sheet.cell(row=i, column=1).number_format = '0.00%'

            self.workbook.save(self.current_filepath)
            
            return {
                "success": True,
                "message": f"ROI and Net Profit formulas added to sheet '{sheet_name}'. Formulas reference '{data_source_sheet}'.",
                "dashboard_sheet": sheet_name,
                "data_source_sheet": data_source_sheet,
                "roi_column": roi_col_name,
                "net_profit_column": net_profit_col_name
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to set up ROI calculations: {str(e)}"
            }
    
    def create_dynamic_charts(self, 
                              chart_data_sheet: str = "FinancialData", 
                              dashboard_sheet: str = "Charts", 
                              chart_type: str = "bar", 
                              values_cols: List[str] = ["Revenue", "Investment"],
                              categories_col: Optional[str] = "Date",
                              chart_title: str = "Financial Performance") -> Dict[str, Any]:
        """
        Creates charts on a specified dashboard sheet based on data from another sheet.
        The charts are 'dynamic' in the sense that openpyxl embeds data and references,
        so Excel will automatically update them if the source data range is adjusted
        or if the data itself changes.
        """
        try:
            if self.workbook is None:
                return {"success": False, "error": "No workbook loaded. Please create or load a spreadsheet first."}
            
            if chart_data_sheet not in self.workbook.sheetnames:
                return {"success": False, "error": f"Chart data source sheet '{chart_data_sheet}' not found."}
            
            data_sheet = self.workbook[chart_data_sheet]
            
            # Create or get the dashboard sheet for charts
            if dashboard_sheet not in self.workbook.sheetnames:
                chart_sheet = self.workbook.create_sheet(dashboard_sheet)
            else:
                chart_sheet = self.workbook[dashboard_sheet]

            # Find column indices for values and categories
            header_row = [cell.value for cell in data_sheet[1]]
            
            value_col_indices = []
            for col_name in values_cols:
                try:
                    value_col_indices.append(header_row.index(col_name) + 1)
                except ValueError:
                    return {"success": False, "error": f"Value column '{col_name}' not found in '{chart_data_sheet}'."}
            
            categories_col_idx = None
            if categories_col:
                try:
                    categories_col_idx = header_row.index(categories_col) + 1
                except ValueError:
                    return {"success": False, "error": f"Categories column '{categories_col}' not found in '{chart_data_sheet}'."}

            num_data_rows = data_sheet.max_row - 1 # Exclude header
            if num_data_rows <= 0:
                return {"success": False, "error": f"No data found in '{chart_data_sheet}' to create charts."}

            # Define data range for the chart
            # Data range for values. Includes header row for series names.
            values_range_start_col_letter = chr(ord("A") + min(value_col_indices) - 1)
            values_range_end_col_letter = chr(ord("A") + max(value_col_indices) - 1)
            values_range = f"'{chart_data_sheet}'!${values_range_start_col_letter}2:${values_range_end_col_letter}{num_data_rows + 1}"
            
            # Data range for categories (if specified)
            categories_range = None
            if categories_col_idx:
                categories_range = f"'{chart_data_sheet}'!${chr(ord('A') + categories_col_idx - 1)}2:'{chart_data_sheet}'!${chr(ord('A') + categories_col_idx - 1)}{num_data_rows + 1}"

            # Create chart object based on type
            if chart_type.lower() == "bar":
                chart = BarChart()
                chart.title = chart_title
                chart.style = 10 # Example style
                chart.y_axis.title = "Amount"
                if categories_col:
                    chart.x_axis.title = categories_col
            # Add more chart types here as needed (e.g., LineChart, PieChart)
            else:
                return {"success": False, "error": f"Unsupported chart type: {chart_type}. Supported types: 'bar'."}
            
            # Add data series
            for col_index in value_col_indices:
                series_data = Reference(data_sheet, min_col=col_index, min_row=2, max_row=num_data_rows + 1)
                # Extract series title from header
                series_title = data_sheet.cell(row=1, column=col_index).value
                
                series = Reference(data_sheet, min_col=col_index, min_row=1, max_row=1) # Reference for series title
                
                chart.add_data(series_data, titles_from_data=False)
                chart.set_categories(Reference(data_sheet, min_col=categories_col_idx, min_row=2, max_row=num_data_rows + 1))
            
            # Add the chart to the sheet
            chart_sheet.add_chart(chart, "A1") # Position the chart at A1

            self.workbook.save(self.current_filepath)
            
            return {
                "success": True,
                "message": f"Chart '{chart_title}' created on sheet '{dashboard_sheet}'.",
                "chart_type": chart_type,
                "values_columns": values_cols,
                "categories_column": categories_col,
                "dashboard_sheet": dashboard_sheet
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to create dynamic chart: {str(e)}"
            }
    
    def add_pivot_table_instructions(self, 
                                     data_source_sheet: str = "FinancialData", 
                                     pivot_sheet: str = "PivotTables", 
                                     pivot_table_name: str = "SalesPivot",
                                     rows: List[str] = ["Region"],
                                     columns: List[str] = ["Product Category"],
                                     values: List[str] = [{"field": "Sales", "aggfunc": "sum"}],
                                     filters: List[str] = ["Year"]) -> Dict[str, Any]:
        """
        This function DOES NOT create the pivot table programmatically as it requires VBA/Macros
        or complex Excel API interactions not directly supported by openpyxl for dynamic pivot generation.
        Instead, it ADDS INSTRUCTIONS OR A TEMPLATE for the user to manually create the pivot table,
        or to use it as a basis for VBA scripting.
        It prepares the data source and provides a clear indication of where to build the pivot.
        """
        try:
            if self.workbook is None:
                return {"success": False, "error": "No workbook loaded. Please create or load a spreadsheet first."}

            if data_source_sheet not in self.workbook.sheetnames:
                return {"success": False, "error": f"Data source sheet '{data_source_sheet}' not found."}
            
            # Ensure pivot sheet exists
            if pivot_sheet not in self.workbook.sheetnames:
                pivot_ws = self.workbook.create_sheet(pivot_sheet)
            else:
                pivot_ws = self.workbook[pivot_sheet]
                # Clear previous content if any to avoid confusion
                pivot_ws.delete_rows(1, pivot_ws.max_row)

            # Add a clear header/instruction
            header_cell = pivot_ws.cell(row=1, column=1)
            header_cell.value = f"Instructions for Pivot Table: '{pivot_table_name}'"
            header_cell.font = Font(bold=True, size=14)
            pivot_ws.merge_cells("A1:E1") # Merge for a prominent title

            pivot_ws.cell(row=3, column=1).value = "To create the pivot table:"
            pivot_ws.cell(row=3, column=1).font = Font(bold=True)

            pivot_ws.cell(row=4, column=2).value = "1. Select the data range in the '" + data_source_sheet + "' sheet."
            pivot_ws.cell(row=5, column=2).value = "2. Go to 'Insert' > 'PivotTable'."
            pivot_ws.cell(row=6, column=2).value = "3. Choose '" + pivot_sheet + "' as the location for the pivot table."
            pivot_ws.cell(row=7, column=2).value = "4. Use the following configuration:"
            
            pivot_ws.cell(row=9, column=3).value = "Pivot Table Name:"
            pivot_ws.cell(row=9, column=4).value = pivot_table_name
            
            pivot_ws.cell(row=10, column=3).value = "Data Source:"
            pivot_ws.cell(row=10, column=4).value = data_source_sheet

            pivot_ws.cell(row=12, column=3).value = "Rows:"
            for i, row_field in enumerate(rows):
                pivot_ws.cell(row=13 + i, column=4).value = row_field
            
            pivot_ws.cell(row=12, column=5).value = "Columns:"
            for i, col_field in enumerate(columns):
                pivot_ws.cell(row=13 + i, column=6).value = col_field
            
            pivot_ws.cell(row=12, column=7).value = "Values:"
            for i, val_config in enumerate(values):
                pivot_ws.cell(row=13 + i, column=8).value = f"{val_config['field']} ({val_config.get('aggfunc', 'sum')})"

            pivot_ws.cell(row=12, column=9).value = "Filters:"
            for i, filter_field in enumerate(filters):
                pivot_ws.cell(row=13 + i, column=10).value = filter_field

            # Add a note about automation
            pivot_ws.cell(row=20, column=1).value = "Note: For full automation, consider using Excel VBA/Macros to create pivot tables programmatically, or use this structure as a template."
            pivot_ws.cell(row=20, column=1).font = Font(italic=True, color="FF808080") # Grey italic text

            self.workbook.save(self.current_filepath)

            return {
                "success": True,
                "message": f"Instructions for creating pivot table '{pivot_table_name}' added to sheet '{pivot_sheet}'.",
                "pivot_sheet": pivot_sheet,
                "data_source_sheet": data_source_sheet,
                "pivot_config": {
                    "rows": rows,
                    "columns": columns,
                    "values": values,
                    "filters": filters
                }
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to add pivot table instructions: {str(e)}"
            }

