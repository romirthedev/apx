
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font
from typing import Dict, Any, List, Optional
import os
from datetime import datetime

class SpreadsheetTool:
    """Tool for creating and manipulating spreadsheets with advanced Excel features."""

    def __init__(self):
        self.current_file_path: Optional[str] = None
        self.current_workbook: Optional[openpyxl.workbook.workbook.Workbook] = None
        self.current_sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        self.data_frame: Optional[pd.DataFrame] = None

    def _get_or_create_workbook(self, filepath: str, sheet_name: str = "Sheet1") -> tuple[openpyxl.workbook.workbook.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
        """Gets the existing workbook and sheet, or creates them if they don't exist."""
        if self.current_workbook and self.current_sheet and self.current_file_path == filepath:
            return self.current_workbook, self.current_sheet

        if os.path.exists(filepath):
            try:
                workbook = openpyxl.load_workbook(filepath)
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(sheet_name)
                self.current_file_path = filepath
                self.current_workbook = workbook
                self.current_sheet = sheet
                # Attempt to load data from existing sheet if possible
                try:
                    self.data_frame = pd.read_excel(filepath, sheet_name=sheet_name)
                except Exception: # If sheet is not a valid dataframe, reset data_frame
                    self.data_frame = None
                return workbook, sheet
            except Exception as e:
                raise IOError(f"Could not load existing Excel file at {filepath}: {e}")
        else:
            workbook = openpyxl.Workbook()
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
            self.current_file_path = filepath
            self.current_workbook = workbook
            self.current_sheet = sheet
            self.data_frame = None # Ensure data_frame is clean for a new file
            return workbook, sheet

    def _save_workbook(self) -> Dict[str, Any]:
        """Saves the current workbook."""
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No workbook to save. Please create or load one first."}
        try:
            self.current_workbook.save(self.current_file_path)
            return {
                "success": True,
                "filepath": self.current_file_path,
                "message": f"Spreadsheet saved successfully to {self.current_file_path}"
            }
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to save spreadsheet: {str(e)}"}

    def create_spreadsheet(self, data: List[Dict[str, Any]], filename: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """
        Create a new spreadsheet with the provided data and optionally set up advanced features.

        Args:
            data: A list of dictionaries representing the data for the spreadsheet.
            filename: The desired name of the Excel file (e.g., "financial_model.xlsx").
            sheet_name: The name of the sheet to create within the workbook.

        Returns:
            A dictionary indicating the success status, filepath, and a message.
        """
        if not isinstance(data, list) or not all(isinstance(row, dict) for row in data):
            return {"success": False, "message": "Invalid 'data' format. Expected a list of dictionaries."}
        if not filename or not isinstance(filename, str):
            return {"success": False, "message": "Invalid 'filename'. Please provide a valid string filename."}
        if not sheet_name or not isinstance(sheet_name, str):
            return {"success": False, "message": "Invalid 'sheet_name'. Please provide a valid string sheet name."}

        filepath = os.path.expanduser(f"~/Desktop/{filename}")
        try:
            workbook, sheet = self._get_or_create_workbook(filepath, sheet_name)

            if data:
                self.data_frame = pd.DataFrame(data)
                # Clear existing content if we are truly creating a new file or overwriting
                if sheet.max_row > 0 or sheet.max_column > 0:
                    sheet.delete_rows(1, sheet.max_row)
                    sheet.delete_cols(1, sheet.max_column)
                
                for r_idx, r in enumerate(dataframe_to_rows(self.data_frame, index=False, header=True), 1):
                    for c_idx, value in enumerate(r, 1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)
            else:
                self.data_frame = pd.DataFrame() # Empty dataframe if no data provided

            self.current_file_path = filepath # Ensure current file is set
            self.current_sheet = sheet
            self.current_workbook = workbook

            result = self._save_workbook()
            if result["success"]:
                result["message"] = f"Spreadsheet '{filename}' with sheet '{sheet_name}' created successfully at {filepath}"
            return result

        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to create spreadsheet: {str(e)}"}

    def add_data(self, new_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Add new data rows to the current spreadsheet.

        Args:
            new_data: A list of dictionaries representing the new data rows to add.

        Returns:
            A dictionary indicating success status, rows added, total rows, and a message.
        """
        if not isinstance(new_data, list) or not all(isinstance(row, dict) for row in new_data):
            return {"success": False, "message": "Invalid 'new_data' format. Expected a list of dictionaries."}
        
        if not self.current_file_path or not self.current_workbook or not self.current_sheet:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' first."}

        try:
            if self.data_frame is None:
                # If no DataFrame was loaded initially, create one from existing sheet data
                try:
                    self.data_frame = pd.read_excel(self.current_file_path, sheet_name=self.current_sheet.title)
                except Exception:
                    self.data_frame = pd.DataFrame() # Start with empty if sheet has no data

            new_df = pd.DataFrame(new_data)
            self.data_frame = pd.concat([self.data_frame, new_df], ignore_index=True)

            # Update the Excel sheet with the new combined data
            self.current_sheet.delete_rows(1, self.current_sheet.max_row)
            self.current_sheet.delete_cols(1, self.current_sheet.max_column)
            for r_idx, r in enumerate(dataframe_to_rows(self.data_frame, index=False, header=True), 1):
                for c_idx, value in enumerate(r, 1):
                    self.current_sheet.cell(row=r_idx, column=c_idx, value=value)

            result = self._save_workbook()
            if result["success"]:
                result["rows_added"] = len(new_data)
                result["total_rows"] = len(self.data_frame)
                result["message"] = f"Added {len(new_data)} rows successfully. Total rows: {len(self.data_frame)}."
            return result

        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to add data: {str(e)}"}

    def add_advanced_formula(self, formula: str, cell: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Add an advanced Excel formula to a specific cell.

        Args:
            formula: The Excel formula string (e.g., "=SUM(A1:A10)").
            cell: The target cell address (e.g., "B1").
            sheet_name: The name of the sheet to add the formula to. If None, uses the current sheet.

        Returns:
            A dictionary indicating success status and a message.
        """
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' first."}
        
        if not formula or not isinstance(formula, str):
            return {"success": False, "message": "Invalid 'formula' format. Expected a non-empty string."}
        if not cell or not isinstance(cell, str):
            return {"success": False, "message": "Invalid 'cell' format. Expected a valid cell address (e.g., 'A1')."}

        try:
            target_sheet = self.current_sheet
            if sheet_name:
                if sheet_name in self.current_workbook.sheetnames:
                    target_sheet = self.current_workbook[sheet_name]
                else:
                    return {"success": False, "message": f"Sheet '{sheet_name}' not found in the current workbook."}
            
            if not target_sheet:
                 return {"success": False, "message": "No active sheet found. Please specify a sheet_name or ensure a sheet is active."}

            target_sheet[cell] = formula
            
            result = self._save_workbook()
            if result["success"]:
                result["message"] = f"Formula '{formula}' added to cell {cell} in sheet '{target_sheet.title}'."
            return result

        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to add formula: {str(e)}"}

    def create_pivot_table(self, data_range: str, pivot_table_destination: str, row_fields: List[str], column_fields: List[str], values_field: str, aggregation_function: str = 'sum', sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Create a pivot table on a specified sheet.

        NOTE: Direct creation of complex pivot tables with openpyxl is limited.
              This function outlines a simplified approach for demonstration.
              For truly advanced pivot tables, programmatic generation in Python
              using libraries that interact with Excel's COM object (like `pywin32`)
              or by manually constructing the data and using Excel's pivot table features
              after export are more robust. This implementation simulates the setup.

        Args:
            data_range: The range of data for the pivot table (e.g., "A1:D10").
            pivot_table_destination: The starting cell for the pivot table (e.g., "F1").
            row_fields: List of column names to use as row labels.
            column_fields: List of column names to use as column labels.
            values_field: The column name for the values to aggregate.
            aggregation_function: The aggregation function ('sum', 'count', 'average', etc.).
            sheet_name: The name of the sheet containing the data and where the pivot table will be placed. If None, uses the current sheet.

        Returns:
            A dictionary indicating success status and a message.
        """
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' first."}
        
        # Basic validation for input parameters
        if not data_range or not isinstance(data_range, str):
            return {"success": False, "message": "Invalid 'data_range'. Expected a string like 'A1:D10'."}
        if not pivot_table_destination or not isinstance(pivot_table_destination, str):
            return {"success": False, "message": "Invalid 'pivot_table_destination'. Expected a cell address like 'F1'."}
        if not isinstance(row_fields, list) or not row_fields:
            return {"success": False, "message": "Invalid 'row_fields'. Expected a non-empty list of column names."}
        if not isinstance(column_fields, list): # Allow empty list for column fields
            return {"success": False, "message": "Invalid 'column_fields'. Expected a list of column names."}
        if not values_field or not isinstance(values_field, str):
            return {"success": False, "message": "Invalid 'values_field'. Expected a column name for aggregation."}
        if aggregation_function not in ['sum', 'count', 'average', 'max', 'min']:
            return {"success": False, "message": f"Unsupported aggregation function: {aggregation_function}. Supported functions are 'sum', 'count', 'average', 'max', 'min'."}

        try:
            target_sheet = self.current_sheet
            if sheet_name:
                if sheet_name in self.current_workbook.sheetnames:
                    target_sheet = self.current_workbook[sheet_name]
                else:
                    return {"success": False, "message": f"Sheet '{sheet_name}' not found in the current workbook."}

            if not target_sheet:
                 return {"success": False, "message": "No active sheet found. Please specify a sheet_name or ensure a sheet is active."}

            # Note: openpyxl does not have direct pivot table creation API.
            # This is a placeholder to indicate where pivot table logic would go.
            # A common approach is to generate the source data and instruct the user
            # to create the pivot table manually, or use more advanced libraries.
            # For a basic representation, we can add a note or potentially a simple chart.

            # Let's add a placeholder text indicating where the pivot table would be
            target_sheet[pivot_table_destination] = "Pivot Table Placeholder"
            target_sheet.cell(row=target_sheet[pivot_table_destination].row + 1, column=target_sheet[pivot_table_destination].column).value = f"Source Data: {data_range}"
            target_sheet.cell(row=target_sheet[pivot_table_destination].row + 2, column=target_sheet[pivot_table_destination].column).value = f"Rows: {', '.join(row_fields)}"
            target_sheet.cell(row=target_sheet[pivot_table_destination].row + 3, column=target_sheet[pivot_table_destination].column).value = f"Columns: {', '.join(column_fields)}"
            target_sheet.cell(row=target_sheet[pivot_table_destination].row + 4, column=target_sheet[pivot_table_destination].column).value = f"Values: {values_field} ({aggregation_function})"
            
            # Attempt to create a simple chart based on aggregated data if data_frame is available
            if self.data_frame is not None and not self.data_frame.empty:
                try:
                    # Prepare data for a simple chart if possible
                    if row_fields and values_field in self.data_frame.columns:
                        agg_col = row_fields[0] # Use the first row field for chart axis
                        
                        if aggregation_function == 'sum':
                            pivot_df = self.data_frame.groupby(agg_col)[values_field].sum().reset_index()
                        elif aggregation_function == 'count':
                            pivot_df = self.data_frame.groupby(agg_col)[values_field].count().reset_index()
                        elif aggregation_function == 'average':
                            pivot_df = self.data_frame.groupby(agg_col)[values_field].mean().reset_index()
                        elif aggregation_function == 'max':
                            pivot_df = self.data_frame.groupby(agg_col)[values_field].max().reset_index()
                        elif aggregation_function == 'min':
                            pivot_df = self.data_frame.groupby(agg_col)[values_field].min().reset_index()
                        else:
                            pivot_df = pd.DataFrame() # Empty if unsupported

                        if not pivot_df.empty and len(pivot_df) > 0:
                            chart_start_row = target_sheet[pivot_table_destination].row + 6
                            chart_start_col = target_sheet[pivot_table_destination].column

                            # Write aggregated data to a temporary area for charting
                            temp_chart_sheet_name = f"{target_sheet.title}_ChartData"
                            if temp_chart_sheet_name not in self.current_workbook.sheetnames:
                                chart_data_sheet = self.current_workbook.create_sheet(temp_chart_sheet_name)
                            else:
                                chart_data_sheet = self.current_workbook[temp_chart_sheet_name]
                                chart_data_sheet.delete_rows(1, chart_data_sheet.max_row)
                                chart_data_sheet.delete_cols(1, chart_data_sheet.max_column)
                            
                            for r_idx, r in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), 1):
                                for c_idx, value in enumerate(r, 1):
                                    chart_data_sheet.cell(row=r_idx, column=c_idx, value=value)

                            # Create a Bar Chart
                            chart = BarChart()
                            chart.title = f"{values_field} by {agg_col}"
                            chart.x_axis.title = agg_col
                            chart.y_axis.title = values_field

                            labels = Reference(chart_data_sheet, min_col=1, min_row=2, max_row=len(pivot_df))
                            data = Reference(chart_data_sheet, min_col=2, min_row=1, max_row=len(pivot_df))
                            chart.add_data(data, titles_from_data=True)
                            chart.set_categories(labels)

                            target_sheet.add_chart(chart, f"{get_column_letter(chart_start_col)}{chart_start_row}")
                            
                            result = self._save_workbook()
                            if result["success"]:
                                result["message"] = (f"Pivot table placeholder added to '{target_sheet.title}'. "
                                                     f"A basic chart has also been generated as '{temp_chart_sheet_name}' is not directly possible in openpyxl. "
                                                     f"Please create the actual pivot table manually in Excel.")
                            return result
                        else:
                            return {"success": False, "message": "Could not generate data for chart from provided fields."}

                    else:
                         result = self._save_workbook()
                         if result["success"]:
                             result["message"] = (f"Pivot table placeholder added to '{target_sheet.title}'. "
                                                  f"Direct pivot table creation with openpyxl is limited. Please create it manually in Excel.")
                         return result
                except Exception as chart_e:
                    print(f"Warning: Could not generate chart for pivot table: {chart_e}")
                    result = self._save_workbook()
                    if result["success"]:
                         result["message"] = (f"Pivot table placeholder added to '{target_sheet.title}'. "
                                              f"Direct pivot table creation with openpyxl is limited. Please create it manually in Excel.")
                    return result
            else:
                result = self._save_workbook()
                if result["success"]:
                    result["message"] = (f"Pivot table placeholder added to '{target_sheet.title}'. "
                                         f"Direct pivot table creation with openpyxl is limited. Please create it manually in Excel.")
                return result

        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to create pivot table: {str(e)}"}

    def apply_conditional_formatting(self, cell_range: str, rule_type: str, operator: str, formula_or_value: Any, sheet_name: Optional[str] = None, fill_color: str = "FFFF00", font_color: str = "000000") -> Dict[str, Any]:
        """
        Apply conditional formatting to a range of cells.

        Args:
            cell_range: The range of cells to apply formatting to (e.g., "A1:A10").
            rule_type: Type of rule ('cellIs', 'formula', 'blanks', 'noBlanks', 'errors', 'noErrors', etc.).
            operator: The operator for 'cellIs' rule (e.g., 'greaterThan', 'lessThan', 'equal', 'between', 'notBetween').
            formula_or_value: The value or formula to compare against.
            sheet_name: The name of the sheet to apply formatting to. If None, uses the current sheet.
            fill_color: Hex color code for the fill (e.g., "FFFF00" for yellow).
            font_color: Hex color code for the font.

        Returns:
            A dictionary indicating success status and a message.
        """
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' first."}
        
        if not cell_range or not isinstance(cell_range, str):
            return {"success": False, "message": "Invalid 'cell_range'. Expected a string like 'A1:A10'."}
        if not rule_type or not isinstance(rule_type, str):
            return {"success": False, "message": "Invalid 'rule_type'. Expected a string representing the rule type."}

        try:
            target_sheet = self.current_sheet
            if sheet_name:
                if sheet_name in self.current_workbook.sheetnames:
                    target_sheet = self.current_workbook[sheet_name]
                else:
                    return {"success": False, "message": f"Sheet '{sheet_name}' not found in the current workbook."}
            
            if not target_sheet:
                 return {"success": False, "message": "No active sheet found. Please specify a sheet_name or ensure a sheet is active."}

            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            font = Font(color=font_color)

            rule = None
            if rule_type == 'cellIs':
                if not operator or not isinstance(operator, str):
                    return {"success": False, "message": "Operator is required for 'cellIs' rule type."}
                rule = CellIsRule(operator=operator, formula=[formula_or_value], stopIfTrue=False, fill=fill, font=font)
            elif rule_type == 'formula':
                rule = CellIsRule(formula=[formula_or_value], stopIfTrue=False, fill=fill, font=font)
            elif rule_type in ['blanks', 'noBlanks', 'errors', 'noErrors']:
                 rule = CellIsRule(operator=rule_type, stopIfTrue=False, fill=fill, font=font) # Simplified for basic types
            else:
                return {"success": False, "message": f"Unsupported rule_type: {rule_type}. Supported types include 'cellIs', 'formula'."}

            if rule:
                target_sheet.conditional_formatting.add(cell_range, rule)
            else:
                return {"success": False, "message": "Failed to create a valid conditional formatting rule."}

            result = self._save_workbook()
            if result["success"]:
                result["message"] = f"Conditional formatting applied to range {cell_range} in sheet '{target_sheet.title}'."
            return result

        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to apply conditional formatting: {str(e)}"}

    def load_spreadsheet(self, filename: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """
        Load an existing spreadsheet and its data.

        Args:
            filename: The name of the Excel file to load.
            sheet_name: The name of the sheet to load data from.

        Returns:
            A dictionary indicating success status, filepath, rows, columns, and a message.
        """
        if not filename or not isinstance(filename, str):
            return {"success": False, "message": "Invalid 'filename'. Please provide a valid string filename."}
        if not sheet_name or not isinstance(sheet_name, str):
            return {"success": False, "message": "Invalid 'sheet_name'. Please provide a valid string sheet name."}

        filepath = os.path.expanduser(f"~/Desktop/{filename}")

        if not os.path.exists(filepath):
            return {"success": False, "message": f"File not found at {filepath}"}

        try:
            workbook = openpyxl.load_workbook(filepath)
            if sheet_name not in workbook.sheetnames:
                return {"success": False, "message": f"Sheet '{sheet_name}' not found in {filename}."}
            
            sheet = workbook[sheet_name]
            
            # Attempt to load data into DataFrame
            try:
                self.data_frame = pd.read_excel(filepath, sheet_name=sheet_name)
                rows = len(self.data_frame)
                columns = len(self.data_frame.columns)
            except Exception:
                self.data_frame = None # If sheet has no data or is not a dataframe
                rows = sheet.max_row if sheet.max_row > 0 else 0
                columns = sheet.max_column if sheet.max_column > 0 else 0

            self.current_file_path = filepath
            self.current_workbook = workbook
            self.current_sheet = sheet

            return {
                "success": True,
                "filepath": filepath,
                "sheet_name": sheet_name,
                "rows": rows,
                "columns": columns,
                "message": f"Spreadsheet '{filename}' (sheet '{sheet_name}') loaded successfully."
            }
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to load spreadsheet: {str(e)}"}

    def save_spreadsheet(self) -> Dict[str, Any]:
        """
        Save the current spreadsheet to its associated file path.

        Returns:
            A dictionary indicating success status and a message.
        """
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' or 'load_spreadsheet' first."}
        
        return self._save_workbook()

    def get_sheet_data(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Retrieve data from a specific sheet as a list of dictionaries.

        Args:
            sheet_name: The name of the sheet to retrieve data from. If None, uses the current sheet.

        Returns:
            A dictionary containing the data, or an error message.
        """
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' or 'load_spreadsheet' first."}

        target_sheet = self.current_sheet
        if sheet_name:
            if sheet_name in self.current_workbook.sheetnames:
                target_sheet = self.current_workbook[sheet_name]
            else:
                return {"success": False, "message": f"Sheet '{sheet_name}' not found in the current workbook."}
        
        if not target_sheet:
            return {"success": False, "message": "No active sheet found. Please specify a sheet_name or ensure a sheet is active."}

        try:
            # Prefer reading from the DataFrame if it exists and matches the sheet
            if self.data_frame is not None and self.current_sheet and sheet_name == self.current_sheet.title or sheet_name is None:
                return {
                    "success": True,
                    "data": self.data_frame.to_dict('records'),
                    "columns": self.data_frame.columns.tolist(),
                    "message": f"Data retrieved from DataFrame for sheet '{target_sheet.title}'."
                }
            else:
                # Fallback to reading directly from the sheet if no DataFrame or sheet mismatch
                data = []
                headers = [cell.value for cell in target_sheet[1]] # Assuming headers are in the first row
                if not headers or all(h is None for h in headers): # If no headers found, try to read all cells
                     for row in target_sheet.iter_rows(min_row=1):
                         row_data = [cell.value for cell in row]
                         if any(cell.value is not None for cell in row): # Only add non-empty rows
                             data.append(row_data)
                     return {
                        "success": True,
                        "data": data,
                        "message": f"Data retrieved from sheet '{target_sheet.title}' (no headers detected)."
                     }
                else:
                    for row_idx, row in enumerate(target_sheet.iter_rows(min_row=2), start=2): # Start from the second row
                        row_dict = {}
                        for col_idx, cell in enumerate(row):
                            if col_idx < len(headers): # Ensure we don't go out of bounds for headers
                                header = headers[col_idx]
                                if header is not None: # Only add if header is not None
                                     row_dict[header] = cell.value
                        if row_dict: # Only add if the dictionary is not empty
                             data.append(row_dict)

                    return {
                        "success": True,
                        "data": data,
                        "columns": [h for h in headers if h is not None],
                        "message": f"Data retrieved from sheet '{target_sheet.title}'."
                    }

        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to retrieve data from sheet: {str(e)}"}

    def add_chart(self, chart_type: str, data_range: str, x_axis_range: str, chart_title: str, output_cell: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Add a chart to the spreadsheet.

        Args:
            chart_type: Type of chart (e.g., 'bar', 'line', 'pie').
            data_range: The range of data for the chart values (e.g., "B2:B10").
            x_axis_range: The range for the x-axis labels (e.g., "A2:A10").
            chart_title: The title of the chart.
            output_cell: The top-left cell where the chart will be placed (e.g., "D2").
            sheet_name: The name of the sheet to add the chart to. If None, uses the current sheet.

        Returns:
            A dictionary indicating success status and a message.
        """
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' or 'load_spreadsheet' first."}
        
        # Input validation
        valid_chart_types = ['bar', 'line', 'pie']
        if chart_type.lower() not in valid_chart_types:
            return {"success": False, "message": f"Invalid chart_type. Supported types are: {', '.join(valid_chart_types)}."}
        if not data_range or not isinstance(data_range, str):
            return {"success": False, "message": "Invalid 'data_range'. Expected a string like 'B2:B10'."}
        if not x_axis_range or not isinstance(x_axis_range, str):
            return {"success": False, "message": "Invalid 'x_axis_range'. Expected a string like 'A2:A10'."}
        if not chart_title or not isinstance(chart_title, str):
            return {"success": False, "message": "Invalid 'chart_title'. Expected a non-empty string."}
        if not output_cell or not isinstance(output_cell, str):
            return {"success": False, "message": "Invalid 'output_cell'. Expected a cell address like 'D2'."}

        try:
            target_sheet = self.current_sheet
            if sheet_name:
                if sheet_name in self.current_workbook.sheetnames:
                    target_sheet = self.current_workbook[sheet_name]
                else:
                    return {"success": False, "message": f"Sheet '{sheet_name}' not found in the current workbook."}
            
            if not target_sheet:
                 return {"success": False, "message": "No active sheet found. Please specify a sheet_name or ensure a sheet is active."}

            chart = None
            if chart_type.lower() == 'bar':
                chart = BarChart()
            elif chart_type.lower() == 'line':
                chart = LineChart()
            elif chart_type.lower() == 'pie':
                chart = PieChart()
            
            if not chart:
                return {"success": False, "message": "Chart object could not be created."}

            # Parse ranges to create Reference objects
            try:
                data_ref = Reference(target_sheet, *openpyxl.utils.cell.cell_range(data_range))
                x_ref = Reference(target_sheet, *openpyxl.utils.cell.cell_range(x_axis_range))
            except Exception as range_parse_error:
                return {"success": False, "message": f"Error parsing cell ranges: {range_parse_error}"}

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(x_ref)
            chart.title = chart_title

            if chart_type.lower() == 'bar':
                chart.x_axis.title = "Category" # Default labels
                chart.y_axis.title = "Value"
            elif chart_type.lower() == 'line':
                chart.x_axis.title = "Category"
                chart.y_axis.title = "Value"

            target_sheet.add_chart(chart, output_cell)

            result = self._save_workbook()
            if result["success"]:
                result["message"] = f"Chart '{chart_title}' ({chart_type.capitalize()}) added to sheet '{target_sheet.title}' at cell {output_cell}."
            return result

        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to add chart: {str(e)}"}

    def remove_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """
        Remove a sheet from the current workbook.

        Args:
            sheet_name: The name of the sheet to remove.

        Returns:
            A dictionary indicating success status and a message.
        """
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' or 'load_spreadsheet' first."}
        
        if not sheet_name or not isinstance(sheet_name, str):
            return {"success": False, "message": "Invalid 'sheet_name'. Please provide a valid string sheet name."}

        if sheet_name not in self.current_workbook.sheetnames:
            return {"success": False, "message": f"Sheet '{sheet_name}' not found in the current workbook."}
        
        if self.current_sheet and self.current_sheet.title == sheet_name:
            # If removing the active sheet, unset it
            self.current_sheet = None
            self.data_frame = None

        try:
            del self.current_workbook[sheet_name]
            result = self._save_workbook()
            if result["success"]:
                result["message"] = f"Sheet '{sheet_name}' removed successfully."
            return result
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to remove sheet: {str(e)}"}

    def set_current_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """
        Set the active sheet for subsequent operations.

        Args:
            sheet_name: The name of the sheet to set as active.

        Returns:
            A dictionary indicating success status and a message.
        """
        if not self.current_file_path or not self.current_workbook:
            return {"success": False, "message": "No spreadsheet is currently loaded or created. Please use 'create_spreadsheet' or 'load_spreadsheet' first."}
        
        if not sheet_name or not isinstance(sheet_name, str):
            return {"success": False, "message": "Invalid 'sheet_name'. Please provide a valid string sheet name."}

        if sheet_name not in self.current_workbook.sheetnames:
            return {"success": False, "message": f"Sheet '{sheet_name}' not found in the current workbook."}
        
        try:
            self.current_sheet = self.current_workbook[sheet_name]
            # Try to load data into DataFrame for the new active sheet
            try:
                self.data_frame = pd.read_excel(self.current_file_path, sheet_name=sheet_name)
            except Exception:
                self.data_frame = None # Reset if the sheet is not a valid dataframe
            
            return {"success": True, "message": f"Active sheet set to '{sheet_name}'."}
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to set current sheet: {str(e)}"}
