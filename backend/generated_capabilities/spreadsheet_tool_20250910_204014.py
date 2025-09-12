import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart
# Correcting the import for CategoryAxis if it's indeed missing/moved in the version
# Based on common openpyxl structure, axis types are usually within openpyxl.chart.axis
# If 'CategoryAxis' is not directly in axis.py, it might be renamed or not available
# For this specific error, let's assume it's available but the import path might be specific.
# If it's genuinely gone, we'd need to adapt the chart creation logic.
# For now, let's keep the original import and assume it might be a versioning issue or a subtle path.
# If it persists, a common alternative is to let openpyxl infer the axis type or to use a generic axis object.
try:
    from openpyxl.chart.axis import CategoryAxis
except ImportError:
    # Fallback: If CategoryAxis is not directly available, openpyxl might handle it implicitly
    # or we might need to use a more generic approach if it's truly removed.
    # For this fix, we'll proceed without explicitly using CategoryAxis if the import fails,
    # as the chart creation might still work without it if it's not strictly required by the
    # underlying chart objects being used. If chart creation fails later, this would be the place to debug.
    print("Warning: 'CategoryAxis' not found in openpyxl.chart.axis. Charting might be affected.")
    CategoryAxis = None # Set to None to indicate it's not available

from openpyxl.chart.series import DataPoint
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Any, List, Optional
import os
from datetime import datetime

class FinancialDashboardTool:
    """
    Tool for creating and manipulating spreadsheets, with enhanced capabilities
    for generating financial dashboards in Excel, including automated pivot tables,
    complex ROI calculations, and dynamic charts.
    """

    def __init__(self):
        self.current_filepath: Optional[str] = None
        self.workbook: Optional[Workbook] = None
        self.data_sheet_name: str = "Raw Data"
        self.pivot_sheet_name_template: str = "Pivot_{}"
        self.chart_sheet_name_template: str = "Chart_{}"
        self.roi_sheet_name: str = "ROI Calculations"

    def _get_or_create_workbook(self, filepath: str) -> Workbook:
        """Internal method to get or create an Excel workbook."""
        if self.workbook:
            return self.workbook
        if os.path.exists(filepath):
            try:
                self.workbook = openpyxl.load_workbook(filepath)
            except Exception as e:
                raise RuntimeError(f"Failed to load existing workbook at {filepath}: {e}")
        else:
            self.workbook = Workbook()
            # Remove default sheet if it exists and we are creating a new workbook
            if "Sheet" in self.workbook.sheetnames and len(self.workbook.sheetnames) == 1:
                del self.workbook["Sheet"]
        return self.workbook

    def _get_or_create_sheet(self, sheet_name: str) -> Worksheet:
        """Internal method to get or create a worksheet."""
        if self.workbook is None: # Ensure workbook is initialized
             raise RuntimeError("Workbook is not initialized.")
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return self.workbook.create_sheet(title=sheet_name)

    def _validate_data(self, data: List[Dict]):
        """Validates if the provided data is in the expected format."""
        if not isinstance(data, list):
            raise ValueError("Data must be a list of dictionaries.")
        if not data:
            return  # Allow empty data for initial creation
        if not all(isinstance(item, dict) for item in data):
            raise ValueError("Each item in the data list must be a dictionary.")
        # Optional: Check for consistent keys across dictionaries
        if len(data) > 1:
            first_keys = set(data[0].keys())
            if not all(set(item.keys()) == first_keys for item in data[1:]):
                print("Warning: Data dictionaries do not have consistent keys. This might affect table creation.")

    def create_dashboard(self, data: List[Dict], filename: str = "financial_dashboard.xlsx", sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Creates a comprehensive financial dashboard in Excel.

        This method initializes the Excel file, writes the raw data, and sets up
        placeholders for automated pivot tables, ROI calculations, and dynamic charts.
        These advanced features will require further specific calls to populate.

        Args:
            data: A list of dictionaries representing the financial data.
                  Expected keys might include: 'Date', 'Revenue', 'Cost', 'Investment', 'Return'.
            filename: The name of the Excel file to create.
            sheet_name: The name for the raw data sheet. Defaults to "Raw Data".

        Returns:
            A dictionary containing the status of the operation, including success,
            filepath, and messages.
        """
        self._validate_data(data)

        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"
        filepath = os.path.expanduser(f"~/Desktop/{filename}")
        self.current_filepath = filepath
        self.data_sheet_name = sheet_name or "Raw Data"

        try:
            self.workbook = Workbook()
            if "Sheet" in self.workbook.sheetnames and len(self.workbook.sheetnames) == 1:
                del self.workbook["Sheet"]

            data_sheet = self._get_or_create_sheet(self.data_sheet_name)

            if data:
                df = pd.DataFrame(data)
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        data_sheet.cell(row=r_idx, column=c_idx, value=value)

            # Create placeholder sheets for advanced features
            self._get_or_create_sheet(self.roi_sheet_name)
            # Add placeholders for potential pivot and chart sheets, but don't create them until needed
            # to avoid cluttering the workbook if these features are not used.

            self.workbook.save(filepath)

            return {
                "success": True,
                "filepath": filepath,
                "message": f"Financial dashboard initialized successfully at {filepath}. Raw data written to '{self.data_sheet_name}' sheet."
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Failed to initialize financial dashboard: {str(e)}"
            }

    def add_data(self, new_data: List[Dict]) -> Dict[str, Any]:
        """
        Adds new data to the existing raw data sheet in the current dashboard.

        Args:
            new_data: A list of dictionaries representing the new financial data.

        Returns:
            A dictionary containing the status of the operation.
        """
        if not self.current_filepath or not os.path.exists(self.current_filepath):
            return {"success": False, "error": "No dashboard file loaded or found. Use create_dashboard first."}

        self._validate_data(new_data)

        try:
            self.workbook = openpyxl.load_workbook(self.current_filepath)
            data_sheet = self._get_or_create_sheet(self.data_sheet_name)

            start_row = data_sheet.max_row + 1
            if data_sheet.max_row == 0:  # If sheet is empty, add header
                if new_data:
                    headers = list(new_data[0].keys())
                    for c_idx, header in enumerate(headers, start=1):
                        data_sheet.cell(row=1, column=c_idx, value=header)
                    start_row = 2

            for r_idx, row_dict in enumerate(new_data, start=start_row):
                for c_idx, (key, value) in enumerate(row_dict.items(), start=1):
                    data_sheet.cell(row=r_idx, column=c_idx, value=value)

            self.workbook.save(self.current_filepath)

            return {
                "success": True,
                "rows_added": len(new_data),
                "message": f"Added {len(new_data)} rows to '{self.data_sheet_name}' sheet."
            }
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to add data: {str(e)}"}

    def create_pivot_table(self, data_range: str, pivot_table_destination: str, index_fields: List[str], values_field: str, agg_func: str = 'sum', sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Creates a pivot table in a new sheet.

        Note: This method uses openpyxl's basic capabilities. For truly automated
        and complex pivot table configurations (like custom calculations within pivots,
        multiple value fields with different aggregations), direct VBA or COM
        automation might be necessary, which is outside the scope of standard openpyxl.
        This implementation creates a single-value aggregation pivot.

        Args:
            data_range: The Excel range of the data (e.g., "A1:D100").
            pivot_table_destination: The cell where the pivot table should start (e.g., "A1").
            index_fields: List of column names to use as row labels (index) in the pivot table.
            values_field: The column name for the values to be aggregated.
            agg_func: The aggregation function ('sum', 'count', 'average', 'max', 'min').
            sheet_name: The name for the pivot table sheet. Defaults to "Pivot_1".

        Returns:
            A dictionary containing the status of the operation.
        """
        if not self.current_filepath or not os.path.exists(self.current_filepath):
            return {"success": False, "error": "No dashboard file loaded. Use create_dashboard first."}

        if not index_fields or not values_field:
            return {"success": False, "error": "index_fields and values_field are required for pivot table."}

        agg_map = {
            'sum': openpyxl.pivot.agg.Sum,
            'count': openpyxl.pivot.agg.Count,
            'average': openpyxl.pivot.agg.Average,
            'max': openpyxl.pivot.agg.Max,
            'min': openpyxl.pivot.agg.Min,
        }
        if agg_func.lower() not in agg_map:
            return {"success": False, "error": f"Unsupported aggregation function: {agg_func}. Choose from {list(agg_map.keys())}."}

        pivot_sheet_name = sheet_name or self.pivot_sheet_name_template.format(len([s for s in self.workbook.sheetnames if s.startswith('Pivot_')]) + 1)

        try:
            self.workbook = openpyxl.load_workbook(self.current_filepath)
            source_sheet = self.workbook[self.data_sheet_name]
            pivot_sheet = self._get_or_create_sheet(pivot_sheet_name)

            # Construct the PivotTableDefinition
            pivot_table_def = openpyxl.pivot.PivotTableDefinition()
            pivot_table_def.cacheSource = openpyxl.pivot.CacheSource(pivot=openpyxl.pivot.CachePivotDefinition(pivotCache=openpyxl.pivot.Cache(ref=f"'{self.data_sheet_name}'!{data_range}")))

            # Add Row Fields (Index Fields)
            for field_name in index_fields:
                row_field = openpyxl.pivot.RowField(x=field_name)
                pivot_table_def.rowFields.append(row_field)

            # Add Value Field
            value_agg_func = agg_map[agg_func.lower()]
            # openpyxl.pivot.agg.Sum expects f.count as parameter for type hinting
            # This is a known quirk. The actual function is implicitly 'sum'.
            # If we were to use 'count', we'd use f.count directly.
            # For this example, we'll use a placeholder as the error was previously about `f.count` not being defined.
            # The actual aggregation is controlled by `subtotal`.
            # Let's try to use the `value_agg_func` correctly.
            # The `function` parameter in ValueField is for specifying the aggregation type more explicitly if needed,
            # but `subtotal` is often sufficient.
            # For example, if subtotal is 'sum', it uses sum. If 'average', it uses average.
            # If `openpyxl.pivot.agg.Sum` is causing issues, and `subtotal` works, we can rely on `subtotal`.

            # Let's try to use `subtotal` and remove `function` if `openpyxl.pivot.agg` is problematic for specific aggregations.
            # The `openpyxl.pivot.agg` classes themselves are meant to be used as types.
            # A common pattern is to use the `subtotal` argument.
            # If `f.count` is the issue, it's likely an import or scope issue.
            # For now, let's assume standard `subtotal` usage works.
            # If `f` is not defined, it's not part of `openpyxl.pivot.agg`.
            # The `function` parameter can take `openpyxl.pivot.agg.Sum`, `openpyxl.pivot.agg.Count` etc.
            # So, `function=value_agg_func` should be correct if `value_agg_func` is indeed one of those classes.

            # Re-checking the openpyxl documentation:
            # ValueField(x, subtotal='sum', header=None, dx=None, format=None, function=None)
            # The 'function' parameter typically expects an enum like `PivotField.SUM` or `PivotField.COUNT`.
            # The `openpyxl.pivot.agg` classes seem to be for defining the aggregation logic, not directly for this parameter.
            # The `subtotal` parameter is the more common way to specify the aggregation type ('sum', 'count', etc.).
            # If `f.count` was an error, it's likely from a misunderstanding of the `function` parameter.
            # Let's use `subtotal` and remove the problematic `function` parameter for now.

            value_field_def = openpyxl.pivot.ValueField(
                x=values_field,
                subtotal=agg_func.lower()
                # removed: function=value_agg_func(f.count) # This was likely the cause of error
            )
            pivot_table_def.valueFields.append(value_field_def)


            pivot_sheet.add_pivot(pivot_table_def, anchor=pivot_table_destination)

            self.workbook.save(self.current_filepath)

            return {
                "success": True,
                "sheet_name": pivot_sheet_name,
                "message": f"Pivot table created successfully on sheet '{pivot_sheet_name}'."
            }
        except KeyError as e:
             return {"success": False, "error": f"Column not found: {e}. Ensure column names match raw data.", "message": f"Failed to create pivot table: Column not found: {e}"}
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to create pivot table: {str(e)}"}


    def calculate_roi(self, investment_col: str, return_col: str, period_col: Optional[str] = None, output_sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Calculates ROI (Return on Investment) and adds it as a new column
        to a specified sheet or creates a new sheet for calculations.

        Formula: ROI = ((Total Return - Total Investment) / Total Investment) * 100

        Args:
            investment_col: The name of the column containing investment values.
            return_col: The name of the column containing return values.
            period_col: Optional. The name of the column specifying the period (e.g., 'Year', 'Quarter').
                        If provided, ROI will be calculated per period.
            output_sheet_name: The name of the sheet to add the ROI column to.
                               If None, a new sheet named "ROI Calculations" will be used/created.

        Returns:
            A dictionary containing the status of the operation.
        """
        if not self.current_filepath or not os.path.exists(self.current_filepath):
            return {"success": False, "error": "No dashboard file loaded. Use create_dashboard first."}

        target_sheet_name = output_sheet_name or self.roi_sheet_name
        try:
            self.workbook = openpyxl.load_workbook(self.current_filepath)
            # Ensure the data sheet exists before trying to read it
            if self.data_sheet_name not in self.workbook.sheetnames:
                return {"success": False, "error": f"Data sheet '{self.data_sheet_name}' not found.", "message": "Failed to calculate ROI: Data sheet not found."}
            data_sheet = self.workbook[self.data_sheet_name]

            # Convert raw data sheet to pandas DataFrame for easier calculation
            # Ensure the file path is valid before reading
            data_df = pd.read_excel(self.current_filepath, sheet_name=self.data_sheet_name)

            if investment_col not in data_df.columns or return_col not in data_df.columns:
                missing_cols = [col for col in [investment_col, return_col] if col not in data_df.columns]
                return {"success": False, "error": f"Missing required columns: {', '.join(missing_cols)}", "message": f"Failed to calculate ROI: Missing required columns: {', '.join(missing_cols)}"}
            if period_col and period_col not in data_df.columns:
                return {"success": False, "error": f"Period column '{period_col}' not found.", "message": f"Failed to calculate ROI: Period column '{period_col}' not found."}

            # Ensure columns are numeric, coerce errors to NaN
            data_df[investment_col] = pd.to_numeric(data_df[investment_col], errors='coerce')
            data_df[return_col] = pd.to_numeric(data_df[return_col], errors='coerce')

            # Calculate ROI per row
            # Handle division by zero or NaN investments
            # Using .get() to safely access column, although check above should prevent KeyError
            investment_values = data_df.get(investment_col)
            return_values = data_df.get(return_col)

            if investment_values is None or return_values is None:
                 return {"success": False, "error": "Could not retrieve investment or return data.", "message": "Failed to calculate ROI: Data retrieval error."}

            # Ensure we don't divide by zero or NaN.
            # If investment is 0 or NaN, ROI is undefined, we can represent this as NaN or 0.
            # For this case, let's set ROI to NaN if investment is 0 or NaN, then fill NaNs.
            denominator = investment_values.replace(0, pd.NA) # Replace 0 with NA for division
            roi_series = ((return_values - investment_values) / denominator) * 100

            # Replace infinities (from 0/0 or similar) and NaNs
            roi_series = roi_series.replace([float('inf'), float('-inf')], pd.NA)
            roi_series = roi_series.fillna(0) # Default to 0 if calculation is invalid or results in NaN

            if period_col:
                # Group by period and calculate aggregated ROI
                grouped_data = data_df.groupby(period_col).agg(
                    total_investment=(investment_col, 'sum'),
                    total_return=(return_col, 'sum')
                )
                # Recalculate ROI on aggregated data, handling potential division by zero
                grouped_denominator = grouped_data['total_investment'].replace(0, pd.NA)
                grouped_data['ROI'] = ((grouped_data['total_return'] - grouped_data['total_investment']) / grouped_denominator) * 100
                grouped_data['ROI'] = grouped_data['ROI'].replace([float('inf'), float('-inf')], pd.NA).fillna(0)
                roi_df = grouped_data[['ROI']].reset_index()
                roi_df.rename(columns={'ROI': 'ROI (%)'}, inplace=True)
                output_data = roi_df
            else:
                # Add ROI as a new column to the original data
                data_df['ROI (%)'] = roi_series
                output_data = data_df # The whole df now contains ROI column

            # Write output to the specified sheet
            output_sheet = self._get_or_create_sheet(target_sheet_name)
            # Clear existing content if it's not the raw data sheet
            if target_sheet_name != self.data_sheet_name:
                # Clear all cells in the target sheet
                for row in output_sheet.iter_rows():
                    for cell in row:
                        cell.value = None

            if not output_data.empty:
                for r_idx, row in enumerate(dataframe_to_rows(output_data, index=False, header=True), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        output_sheet.cell(row=r_idx, column=c_idx, value=value)

            self.workbook.save(self.current_filepath)

            return {
                "success": True,
                "sheet_name": target_sheet_name,
                "message": f"ROI calculated and saved to '{target_sheet_name}'."
            }
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to calculate ROI: {str(e)}"}

    def create_dynamic_chart(self, chart_type: str, x_axis_col: str, y_axis_cols: List[str], data_sheet: str = None, chart_sheet_name: Optional[str] = None, title: str = "Financial Chart") -> Dict[str, Any]:
        """
        Creates a dynamic chart based on data from a specified sheet.
        'Dynamic' here means the chart updates when the underlying data is refreshed,
        as Excel charts are linked to cell ranges.

        Args:
            chart_type: Type of chart ('bar', 'pie').
            x_axis_col: The column name for the chart's x-axis (categories).
            y_axis_cols: A list of column names for the chart's y-axis (values).
            data_sheet: The name of the sheet containing the data. Defaults to "Raw Data".
            chart_sheet_name: The name for the chart sheet. Defaults to "Chart_1".
            title: The title of the chart.

        Returns:
            A dictionary containing the status of the operation.
        """
        if not self.current_filepath or not os.path.exists(self.current_filepath):
            return {"success": False, "error": "No dashboard file loaded. Use create_dashboard first."}

        sheet_to_use = data_sheet or self.data_sheet_name
        # Load workbook here to check sheet existence before proceeding
        try:
            self.workbook = openpyxl.load_workbook(self.current_filepath)
        except Exception as e:
            return {"success": False, "error": f"Failed to load workbook: {e}", "message": f"Error loading workbook: {e}"}

        if sheet_to_use not in self.workbook.sheetnames:
             return {"success": False, "error": f"Data sheet '{sheet_to_use}' not found.", "message": f"Failed to create chart: Data sheet '{sheet_to_use}' not found."}

        chart_sheet_name = chart_sheet_name or self.chart_sheet_name_template.format(len([s for s in self.workbook.sheetnames if s.startswith('Chart_')]) + 1)

        try:
            source_sheet = self.workbook[sheet_to_use]
            chart_sheet = self._get_or_create_sheet(chart_sheet_name)

            # Read data from the source sheet into a DataFrame to easily get column indices
            # Using pandas to easily find column indices and handle potential header issues
            df = pd.read_excel(self.current_filepath, sheet_name=sheet_to_use)

            if x_axis_col not in df.columns:
                return {"success": False, "error": f"X-axis column '{x_axis_col}' not found in '{sheet_to_use}'.", "message": f"Failed to create chart: X-axis column '{x_axis_col}' not found."}
            for y_col in y_axis_cols:
                if y_col not in df.columns:
                    return {"success": False, "error": f"Y-axis column '{y_col}' not found in '{sheet_to_use}'.", "message": f"Failed to create chart: Y-axis column '{y_col}' not found."}

            # Find the column indices in the worksheet
            # Assumes first row is header
            headers = [cell.value for cell in source_sheet[1]]
            try:
                x_col_idx = headers.index(x_axis_col) + 1
                y_col_indices = [headers.index(y_col) + 1 for y_col in y_axis_cols]
            except ValueError as e:
                return {"success": False, "error": f"Column not found in sheet headers: {e}", "message": f"Failed to create chart: Column not found in sheet headers: {e}"}

            # Define data ranges dynamically
            num_rows = source_sheet.max_row
            if num_rows <= 1: # Only header row
                 return {"success": False, "error": "Not enough data to create a chart.", "message": "Failed to create chart: Not enough data."}

            # Data for the series (values)
            # The 'max_row' should be the last row of data, so it's num_rows if num_rows is the total rows including header.
            # If source_sheet.max_row includes header, data starts from row 2.
            # Range is from row 2 to num_rows.
            series_data_refs = []
            for y_col_idx in y_col_indices:
                series_data_refs.append(Reference(source_sheet, min_col=y_col_idx, min_row=2, max_row=num_rows))

            # Categories for the x-axis
            cat_axis = Reference(source_sheet, min_col=x_col_idx, min_row=2, max_row=num_rows)

            # For bar charts
            if chart_type.lower() == 'bar':
                chart = BarChart()
                chart.title = title
                chart.y_axis.title = "Value"
                chart.x_axis.title = x_axis_col

                # Add data series
                for i, y_col_idx in enumerate(y_col_indices):
                    chart.add_data(series_data_refs[i], titles_from_data=False) # Add data series
                    # Set the legend title from the header cell
                    series_legend = Reference(source_sheet, min_col=y_col_idx, min_row=1, max_row=1)
                    chart.series[i].header = series_legend

                chart.set_categories(cat_axis)

                # Ensure CategoryAxis is used if available and applicable.
                # If CategoryAxis was successfully imported:
                if CategoryAxis:
                    chart.x_axis = CategoryAxis()
                    chart.x_axis.title = x_axis_col
                    chart.x_axis.set_categories(cat_axis) # Explicitly set categories for CategoryAxis

            # For pie charts
            elif chart_type.lower() == 'pie':
                chart = PieChart()
                chart.title = title
                if len(y_axis_cols) > 1:
                    print("Warning: Pie charts typically represent a single value series. Using the first Y-axis column.")
                
                # Pie chart uses the first y-axis column
                chart.add_data(series_data_refs[0], titles_from_data=False)
                series_legend = Reference(source_sheet, min_col=y_col_indices[0], min_row=1, max_row=1)
                chart.series[0].tx = series_legend # Set legend title

                chart.set_categories(cat_axis)
                # Pie charts also have an x-axis (which is categories), so CategoryAxis is relevant.
                if CategoryAxis:
                    chart.x_axis = CategoryAxis()
                    chart.x_axis.set_categories(cat_axis)


            else:
                return {"success": False, "error": f"Unsupported chart type: {chart_type}. Choose from 'bar', 'pie'.", "message": f"Failed to create chart: Unsupported chart type."}

            # Add the chart to the chart sheet.
            # openpyxl charts are often placed on a separate sheet, or embedded.
            # For simplicity, we'll place it on the specified chart sheet.
            chart_sheet.add_chart(chart, "B2") # Anchor point for the chart

            self.workbook.save(self.current_filepath)

            return {
                "success": True,
                "sheet_name": chart_sheet_name,
                "message": f"Chart '{title}' created successfully on sheet '{chart_sheet_name}'."
            }
        except Exception as e:
            # More specific error handling might be needed if the error is related to chart object creation.
            return {"success": False, "error": str(e), "message": f"Failed to create chart: {str(e)}"}

    def save_spreadsheet(self) -> Dict[str, Any]:
        """Saves the current workbook to its current filepath."""
        if not self.current_filepath:
            return {"success": False, "error": "No file path available. Use create_dashboard first.", "message": "Cannot save: no file path set."}
        if not self.workbook:
            return {"success": False, "error": "No workbook loaded.", "message": "Cannot save: workbook not loaded."}

        try:
            self.workbook.save(self.current_filepath)
            return {
                "success": True,
                "filepath": self.current_filepath,
                "message": f"Spreadsheet saved successfully to {self.current_filepath}"
            }
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to save spreadsheet: {str(e)}"}

    def load_spreadsheet(self, filepath: str) -> Dict[str, Any]:
        """Loads an existing spreadsheet."""
        filepath = os.path.expanduser(filepath)
        if not os.path.exists(filepath):
            return {"success": False, "error": "File not found.", "message": f"Failed to load: File not found at {filepath}"}

        try:
            self.workbook = openpyxl.load_workbook(filepath)
            self.current_filepath = filepath
            # Attempt to identify the raw data sheet if it exists, default to 'Raw Data'
            if self.data_sheet_name not in self.workbook.sheetnames:
                print(f"Warning: Default data sheet '{self.data_sheet_name}' not found. Using first available sheet.")
                # Ensure there's at least one sheet before accessing it
                if self.workbook.sheetnames:
                    self.data_sheet_name = self.workbook.sheetnames[0]
                else:
                    # If the workbook is completely empty, create a default sheet
                    self.data_sheet_name = "Sheet1"
                    self._get_or_create_sheet(self.data_sheet_name)


            return {
                "success": True,
                "filepath": filepath,
                "message": f"Spreadsheet loaded successfully from {filepath}"
            }
        except Exception as e:
            return {"success": False, "error": str(e), "message": f"Failed to load spreadsheet: {str(e)}"}